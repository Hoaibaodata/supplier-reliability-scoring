"""Generate SupplierReliability.xlsx starter — sheets + headers + Dashboard layout.

User opens → Save As .xlsm → Alt+F11 → Import 5 .bas files → add buttons.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_FONT = Font(name="Segoe UI", size=18, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Segoe UI", size=11, italic=True, color="595959")


def write_headers(ws, headers):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22


def build_dashboard(ws):
    ws["A1"] = "Supplier Reliability Scoring — Dashboard"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    ws["A2"] = "VBA workflow — bấm nút theo thứ tự"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:H2")

    instructions = [
        "",
        "BƯỚC 1 — Generate Sample PO History (200 POs, 10 vendors)",
        "BƯỚC 2 — Compute Scores (aggregate KPIs + composite score + tier)",
        "BƯỚC 3 — Predict Slip Risk (90-day recent vs historical OTD)",
        "BƯỚC 4 — Build Chart (visualize composite scores by tier)",
        "",
        "Hoặc bấm nút 'Run Full Analysis' để chạy hết cùng lúc.",
        "",
        "Xem kết quả ở các sheet: Scorecards, Predictions, và biểu đồ ngay đây.",
    ]
    for i, line in enumerate(instructions, start=4):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = Font(name="Segoe UI", size=11, bold=(i in (5, 6, 7, 8, 10)))
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    ws.column_dimensions["A"].width = 80


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # Dashboard first
    ws_dash = wb.create_sheet("Dashboard")
    build_dashboard(ws_dash)

    # PO_History with headers (empty)
    ws_po = wb.create_sheet("PO_History")
    write_headers(ws_po, [
        "po_id", "vendor_id", "vendor_name", "category",
        "order_date", "requested_delivery_date", "requested_qty",
        "actual_delivery_date", "delivered_qty", "qc_pass_qty", "qc_fail_qty",
        "requested_lt_days", "actual_lt_days", "slip_days",
        "is_otd", "is_in_full", "confirm_lag_days",
    ])
    ws_po["A3"] = "(Bam 'Generate Sample' tren Dashboard de fill 200 PO mau, hoac paste data thuc cua ban vao day.)"

    # Scorecards (empty, will be filled by VBA)
    ws_sc = wb.create_sheet("Scorecards")
    write_headers(ws_sc, [
        "vendor_id", "vendor_name", "num_pos", "otd_rate", "otif_rate",
        "avg_lt_days", "sigma_lt_weeks", "quality_pass_rate", "communication_score",
        "trade_score", "composite_score", "tier",
    ])

    # Predictions (empty)
    ws_pred = wb.create_sheet("Predictions")
    write_headers(ws_pred, [
        "vendor_id", "vendor_name", "tier", "composite_score",
        "historical_otd", "recent_90d_otd", "recent_n",
        "predicted_slip_probability",
    ])

    # Parameters (editable)
    ws_p = wb.create_sheet("Parameters")
    write_headers(ws_p, ["weight_name", "value", "description"])
    params = [
        ("weight_otif", 0.40, "OTIF rate (on-time + in-full)"),
        ("weight_lt_consistency", 0.20, "Lead time CV inverse"),
        ("weight_quality", 0.20, "1 - QC fail rate"),
        ("weight_communication", 0.10, "Communication score from confirm lag"),
        ("weight_trade", 0.10, "Trade relationship duration"),
        ("", "", ""),
        ("tier_A_threshold", 85, "Composite score for A-Preferred"),
        ("tier_B_threshold", 75, "Composite score for B-Standard"),
        ("tier_C_threshold", 60, "Composite score for C-Watch"),
        ("", "", "(Edit values above to retune scoring.)"),
    ]
    for i, (name, val, desc) in enumerate(params, start=2):
        ws_p.cell(row=i, column=1, value=name)
        ws_p.cell(row=i, column=2, value=val)
        ws_p.cell(row=i, column=3, value=desc)
    ws_p.column_dimensions["A"].width = 24
    ws_p.column_dimensions["B"].width = 12
    ws_p.column_dimensions["C"].width = 50

    # Log (empty)
    ws_log = wb.create_sheet("Log")
    write_headers(ws_log, ["Timestamp", "Level", "Source", "Message"])

    out = Path(__file__).resolve().parent / "SupplierReliability.xlsx"
    wb.save(out)
    print(f"[OK] Tao xong: {out}")
    print()
    print("Buoc tiep theo:")
    print("  1. Mo SupplierReliability.xlsx trong Excel")
    print("  2. File > Save As > Save as type: Excel Macro-Enabled Workbook (.xlsm)")
    print("  3. Alt+F11 > File > Import File > chon 5 .bas trong vba_modules/")
    print("  4. Quay lai Excel > File > Save")
    print("  5. Developer tab > Insert Button > Assign macro 'Action_RunFullAnalysis'")
    print("  6. (Optional) Them them cac nut: Action_GenerateSample, Action_ComputeScores,")
    print("     Action_PredictSlip, Action_BuildChart, Action_ResetOutputs")


if __name__ == "__main__":
    main()
